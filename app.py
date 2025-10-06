import subprocess
import sys
from datetime import datetime, timedelta
from decimal import Decimal

import pytz
from apscheduler.schedulers.background import BackgroundScheduler
from waitress import serve

import mean_1h

from flask_bcrypt import Bcrypt
import logging
import os
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    send_file,
    session,
    redirect,
    url_for,
    flash,
    make_response,
    got_request_exception,
)
import pymysql
from pymysql.cursors import DictCursor
import configparser
import openpyxl
from functools import wraps
from contextlib import closing
from typing import Dict, List, Optional, Sequence

import insertMissingDataFromCSV
from wind_utils import calculate_wind_stats
from logging import FileHandler, WARNING
import threading
import time
current_dir = os.path.dirname(os.path.abspath(__file__))


def resource_path(relative_path: str) -> str:
    """Resolve resource paths for both source and PyInstaller builds."""
    candidates = []
    if hasattr(sys, '_MEIPASS'):
        candidates.append(os.path.join(sys._MEIPASS, relative_path))
    if getattr(sys, 'frozen', False):
        candidates.append(os.path.join(os.path.dirname(sys.executable), relative_path))
    candidates.append(os.path.join(current_dir, relative_path))

    for path in candidates:
        if os.path.exists(path):
            return path
    return candidates[-1]
# Create a logger
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

# Create file handler which logs only errors
file_handler = logging.FileHandler('errorlog.log')
file_handler.setLevel(logging.ERROR)

# Create formatter and add it to the file handler
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# Add the file handler to the logger
logger.addHandler(file_handler)

file_handler = FileHandler('errorlog.txt')
file_handler.setLevel(WARNING)
logger.addHandler(file_handler)

# Load configuration from config.ini
with open('config.ini', 'r', encoding='utf-8') as config_file:
    config = configparser.ConfigParser(interpolation=None)
    config.read_file(config_file)

# MySQL Database Configuration
DB_HOST = config.get('SQL', 'host')
DB_PORT = config.getint('SQL', 'port')
DB_USER = config.get('SQL', 'user')
DB_PASSWORD = config.get('SQL', 'password')
DB_NAME = config.get('SQL', 'database')
DB_TABLE = config.get('SQL', 'mean_1hour_table')
DB_TABLE_MIN = config.get('SQL', 'DB_TABLE_MIN')

# Columns
DATE_COLUMN = 'DateRef'
RAW_DATA_COLUMNS = [col.strip() for col in config.get('SQL', 'columns_list').split(',')]
CALCULATED_COLUMNS = [col.strip() for col in config.get('SQL', 'calc_columns').split(',')]
DATA_COLUMNS = RAW_DATA_COLUMNS + CALCULATED_COLUMNS
RAW_UNITS = [col.strip() for col in config.get('SQL', 'columns_units_list').split(',')]
CALCULATED_UNITS = [col.strip() for col in config.get('SQL', 'calc_units_list').split(',')]
DATA_COLUMNS_UNITS = RAW_UNITS + CALCULATED_UNITS
# Pollutant-specific configuration removed; initialize empty lists for compatibility
POLLUTANT_COLUMNS = []
nde_all = []
pollutants_units = []
simple_parameters = []
simple_parameters_units = []
simple_parameters_BG = []


# Names
DATA_COLUMNS_NAMES = DATA_COLUMNS
# Excel template file path
EXCEL_TEMPLATE_PATH = 'template.xlsx'
RAW_BG = [col.strip() for col in config.get('SQL', 'DATA_COLUMNS_BG').split(',')]
CALCULATED_BG = [col.strip() for col in config.get('SQL', 'calc_columns_bg').split(',')]
DATA_COLUMNS_BG = RAW_BG + CALCULATED_BG
# Constants for radiation unit conversion
KWH_PER_M2_FROM_MINUTE = 1 / 60000.0  # Sum of 1-minute W/m² values -> kWh/m²
KWH_PER_M2_FROM_HOUR = 1 / 1000.0  # Sum of 1-hour W/m² values -> kWh/m²
# Variable to store the path of the saved plot image
plot_image_path = 'plot.png'


def _get_config_list(section: str, option: str) -> List[str]:
    if not config.has_section(section) or not config.has_option(section, option):
        return []
    return [item.strip() for item in config.get(section, option).split(',') if item.strip()]


ALARM_TABLE = None
ALARM_DATE_COLUMN = DATE_COLUMN
ALARM_COLUMNS: List[str] = []
ALARM_COLUMN_LABELS: Dict[str, str] = {}
ALARM_COLUMN_DESCRIPTIONS: Dict[str, str] = {}

if config.has_section('ALARMS'):
    ALARM_TABLE = config.get('ALARMS', 'table', fallback=None)
    ALARM_DATE_COLUMN = config.get('ALARMS', 'date_column', fallback=DATE_COLUMN)
    ALARM_COLUMNS = _get_config_list('ALARMS', 'columns')
    alarm_column_names = _get_config_list('ALARMS', 'columns_names')
    alarm_column_descriptions = _get_config_list('ALARMS', 'columns_descriptions')

    for idx, column in enumerate(ALARM_COLUMNS):
        label = alarm_column_names[idx] if idx < len(alarm_column_names) else column
        description = (
            alarm_column_descriptions[idx]
            if idx < len(alarm_column_descriptions)
            else label
        )
        ALARM_COLUMN_LABELS[column] = label
        ALARM_COLUMN_DESCRIPTIONS[column] = description

global my_df
my_df = pd.DataFrame(columns=[DATE_COLUMN] + DATA_COLUMNS)
# Global dataframe to store the required data
df_last_min_values = pd.DataFrame(columns=[DATE_COLUMN] + DATA_COLUMNS)
def add_calculated_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy().sort_index()
    if 'RAIN' in df.columns:
        df['RAIN'] = pd.to_numeric(df['RAIN'], errors='coerce')
    if 'EVAPOR_MINUTE' in df.columns:
        df['EVAPOR_MINUTE'] = pd.to_numeric(df['EVAPOR_MINUTE'], errors='coerce')
        if 'EVAPOR_DAY' not in df.columns:
            df['EVAPOR_DAY'] = np.nan
    return df
# Function to establish a connection to the MySQL database
def get_db_connection():
    return pymysql.connect(host=DB_HOST, user=DB_USER, password=DB_PASSWORD, database=DB_NAME, port=DB_PORT)


def _is_alarm_active(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value) != 0.0
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {'', '0', 'false', 'inactive', 'не'}:
            return False
        return True
    return bool(value)


def get_active_alarms():
    if not ALARM_TABLE or not ALARM_COLUMNS:
        return {'timestamp': None, 'items': []}

    db_connection = None
    cursor = None
    try:
        db_connection = get_db_connection()
        cursor = db_connection.cursor(DictCursor)
        selected_columns = list(dict.fromkeys([ALARM_DATE_COLUMN] + ALARM_COLUMNS))
        columns_clause = ', '.join(selected_columns)
        query = (
            f"SELECT {columns_clause} FROM {ALARM_TABLE} "
            f"ORDER BY {ALARM_DATE_COLUMN} DESC LIMIT 1"
        )
        cursor.execute(query)
        result = cursor.fetchone()
    except Exception as exc:
        logger.error(f"Error fetching alarms: {exc}", exc_info=True)
        return {'timestamp': None, 'items': []}
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if db_connection is not None:
            try:
                db_connection.close()
            except Exception:
                pass

    if not result:
        return {'timestamp': None, 'items': []}

    timestamp = result.get(ALARM_DATE_COLUMN)
    items = []
    for column in ALARM_COLUMNS:
        value = result.get(column)
        if _is_alarm_active(value):
            processed_value = value
            if isinstance(value, Decimal):
                processed_value = float(value)
            elif isinstance(value, np.generic):
                processed_value = value.item()
            elif isinstance(value, datetime):
                processed_value = value.strftime('%Y-%m-%dT%H:%M:%S')
            items.append(
                {
                    'column': column,
                    'name': ALARM_COLUMN_LABELS.get(column, column),
                    'description': ALARM_COLUMN_DESCRIPTIONS.get(column, column),
                    'value': processed_value,
                }
            )

    if isinstance(timestamp, datetime):
        timestamp_str = timestamp.strftime('%Y-%m-%dT%H:%M:%S')
    else:
        timestamp_str = str(timestamp) if timestamp else None

    return {'timestamp': timestamp_str, 'items': items}

#INitialize plot

STATIC_FOLDER = resource_path('static')
template_dir = resource_path('templates')
app = Flask(__name__, static_folder=STATIC_FOLDER, template_folder=template_dir)
#app = Flask(__name__, template_folder=template_dir)
#app = Flask(__name__)
bcrypt = Bcrypt(app)
app.secret_key = 'джасфжфдгсдфг'  # Replace with a secure random key
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_COOKIE_SECURE'] = False  # Allow cookies over HTTP; set to True when using HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Prevent client-side script access to the cookie
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Prevent cross-site requests

# Dummy user store for demonstration
users = {
    'admin': bcrypt.generate_password_hash('password123').decode('utf-8')
}

# Timezone setup (optional, adjust if needed)
TZ = pytz.timezone("Europe/Sofia")  # Adjust timezone as needed

init = 0

def _fetch_last_minute_data(cursor) -> pd.DataFrame:
    query_last_minute = f"""
        SELECT {DATE_COLUMN}, {', '.join(RAW_DATA_COLUMNS)}
        FROM {DB_TABLE_MIN}
        ORDER BY {DATE_COLUMN} DESC LIMIT 1
    """
    cursor.execute(query_last_minute)
    last_minute_data = cursor.fetchall()

    df_last_min_data = pd.DataFrame(
        last_minute_data,
        columns=[DATE_COLUMN] + RAW_DATA_COLUMNS,
    )

    if df_last_min_data.empty:
        return df_last_min_data

    df_last_min_data[RAW_DATA_COLUMNS] = df_last_min_data[RAW_DATA_COLUMNS].apply(
        pd.to_numeric, errors='coerce'
    )
    df_last_min_data[DATE_COLUMN] = pd.to_datetime(df_last_min_data[DATE_COLUMN])
    df_last_min_data.set_index(DATE_COLUMN, inplace=True)
    return add_calculated_columns(df_last_min_data)


def _average_rainfall(cursor, start: datetime, end: datetime) -> float:
    if start >= end:
        return 0.0

    query = (
        f"SELECT AVG(RAIN) FROM {DB_TABLE_MIN} "
        f"WHERE {DATE_COLUMN} >= %s AND {DATE_COLUMN} < %s"
    )

    try:
        cursor.execute(query, (start, end))
        row = cursor.fetchone()
    except Exception as exc:
        logger.error(
            f"Error averaging rainfall between {start} and {end}: {exc}",
            exc_info=True,
        )
        return 0.0

    if not row:
        return 0.0

    value = row[0]
    return float(value) if value is not None else 0.0


def _sum_hourly_rainfall(cursor, start: datetime, end: datetime) -> float:
    if start >= end:
        return 0.0

    query = (
        f"SELECT AVG(RAIN) AS avg_rain FROM {DB_TABLE_MIN} "
        f"WHERE {DATE_COLUMN} >= %s AND {DATE_COLUMN} < %s "
        f"GROUP BY YEAR({DATE_COLUMN}), MONTH({DATE_COLUMN}), DAY({DATE_COLUMN}), HOUR({DATE_COLUMN})"
    )

    try:
        cursor.execute(query, (start, end))
        rows = cursor.fetchall()
    except Exception as exc:
        logger.error(
            f"Error summing hourly rainfall between {start} and {end}: {exc}",
            exc_info=True,
        )
        return 0.0

    total = 0.0
    for row in rows:
        if not row:
            continue
        value = row[0]
        if value is not None:
            total += float(value)
    return total


def _set_rainfall_totals(cursor, df_last_min_data: pd.DataFrame) -> pd.DataFrame:
    if 'RAIN' not in df_last_min_data.columns or df_last_min_data.empty:
        return df_last_min_data

    last_timestamp = df_last_min_data.index[-1]
    if pd.isna(last_timestamp):
        return df_last_min_data

    completed_hour_end = last_timestamp.to_pydatetime().replace(
        minute=0, second=0, microsecond=0
    )
    previous_hour_start = completed_hour_end - timedelta(hours=1)

    rain_hour_total = _average_rainfall(cursor, previous_hour_start, completed_hour_end)
    df_last_min_data.loc[last_timestamp, 'RAIN_HOUR'] = rain_hour_total

    sum_ranges = {
        'RAIN_DAY': completed_hour_end.replace(hour=0, minute=0, second=0, microsecond=0),
        'RAIN_MONTH': completed_hour_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0),
        'RAIN_YEAR': completed_hour_end.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0),
    }

    for column, start_time in sum_ranges.items():
        total = _sum_hourly_rainfall(cursor, start_time, completed_hour_end)
        df_last_min_data.loc[last_timestamp, column] = total

    return df_last_min_data


def _compute_daily_evaporation_mean(
    cursor, timestamp: pd.Timestamp
) -> Optional[float]:
    """Return the average evaporation intensity for the current day so far."""

    if 'EVAPOR_MINUTE' not in RAW_DATA_COLUMNS or pd.isna(timestamp):
        return None

    last_dt = (
        timestamp.to_pydatetime() if isinstance(timestamp, pd.Timestamp) else timestamp
    )
    day_start = last_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    query = (
        f"SELECT AVG(EVAPOR_MINUTE) FROM {DB_TABLE_MIN} "
        f"WHERE {DATE_COLUMN} >= %s AND {DATE_COLUMN} <= %s"
    )

    try:
        cursor.execute(query, (day_start, last_dt))
        result = cursor.fetchone()
    except Exception as exc:
        logger.error(
            f"Error computing daily evaporation average for {timestamp}: {exc}",
            exc_info=True,
        )
        return None

    if not result:
        return None

    value = result[0]
    return float(value) if value is not None else None


def _prepare_last_minute_output(df_last_min_data: pd.DataFrame) -> pd.DataFrame:
    df_last_min_values = df_last_min_data.reindex(columns=DATA_COLUMNS)
    df_last_min_values.reset_index(inplace=True)

    rainfall_cols = [
        col
        for col in ['RAIN', 'RAIN_HOUR', 'RAIN_DAY', 'RAIN_MONTH', 'RAIN_YEAR']
        if col in df_last_min_values.columns
    ]
    numeric_columns = df_last_min_values.select_dtypes(include=[np.number]).columns
    non_rain_cols = [col for col in numeric_columns if col not in rainfall_cols]

    if non_rain_cols:
        df_last_min_values.loc[:, non_rain_cols] = (
            df_last_min_values[non_rain_cols].round(1)
        )

    for col in rainfall_cols:
        df_last_min_values.loc[:, col] = df_last_min_values[col].round(2)

    return df_last_min_values


def update_dataframes():
    global df_last_min_values

    while True:
        try:
            with closing(get_db_connection()) as db_connection:
                with db_connection.cursor() as cursor:
                    df_last_min_data = _fetch_last_minute_data(cursor)

                    if not df_last_min_data.empty:
                        df_last_min_data = _set_rainfall_totals(cursor, df_last_min_data)
                        df_last_min_values = _prepare_last_minute_output(df_last_min_data)

                        if 'EVAPOR_MINUTE' in df_last_min_data.columns:
                            last_timestamp = df_last_min_data.index[-1]
                            daily_evap = _compute_daily_evaporation_mean(
                                cursor, last_timestamp
                            )

                            if (
                                daily_evap is not None
                                and 'EVAPOR_DAY' in df_last_min_values.columns
                            ):
                                df_last_min_values.loc[:, 'EVAPOR_DAY'] = round(
                                    daily_evap, 1
                                )
        except Exception as e:
            logger.error(f"Error updating dataframes: {e}", exc_info=True)

        time.sleep(30)


# Start the update thread
#update_dataframes()
threading.Thread(target=update_dataframes, daemon=True).start()


# Login route
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username in users and bcrypt.check_password_hash(users[username], password):
            session['username'] = username
            flash('Login successful!', 'success')
            return render_template('index.html')
        else:
            error_message = 'Невалидно потребителско име или парола.'
            return render_template('login.html', error_message=error_message)

    return render_template('login.html')


@app.route('/logout', methods=['GET'])
def logout():
    session.clear()  # Clear session data
    flash('You have been logged out.', 'info')  # Optional: Show a message
    return redirect(url_for('login'))  # Redirect to the login page


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        return f(*args, **kwargs)
    return decorated_function


# Route to serve the HTML webpage
@app.route('/')
@login_required
def index():
    return render_template('index.html')  # Serve the main page


@app.route('/graphs')
@login_required
def graphs_page():
    return render_template('graphs.html')


@app.route('/graph_data', methods=['GET'])
@login_required
def graph_data():
    period = request.args.get('period', '24h')
    try:
        db_connection = get_db_connection()
        cursor = db_connection.cursor()
        end_time = datetime.now()

        if period == '24h':
            start_time = end_time - timedelta(hours=24)
            table = DB_TABLE_MIN
        elif period == '30d':
            start_time = end_time - timedelta(days=30)
            table = DB_TABLE
        else:
            start_time = end_time - timedelta(days=365)
            table = DB_TABLE

        query = (
            f"SELECT {DATE_COLUMN}, {', '.join(RAW_DATA_COLUMNS)} FROM {table} "
            f"WHERE {DATE_COLUMN} >= %s AND {DATE_COLUMN} <= %s ORDER BY {DATE_COLUMN} ASC"
        )
        cursor.execute(
            query,
            (
                start_time.strftime('%Y-%m-%d %H:%M:%S'),
                end_time.strftime('%Y-%m-%d %H:%M:%S'),
            ),
        )
        data = cursor.fetchall()
        cursor.close()
        db_connection.close()

        df = pd.DataFrame(data, columns=[DATE_COLUMN] + RAW_DATA_COLUMNS)
        if df.empty:
            return jsonify({})

        df[RAW_DATA_COLUMNS] = df[RAW_DATA_COLUMNS].apply(pd.to_numeric, errors='coerce')
        df[DATE_COLUMN] = pd.to_datetime(df[DATE_COLUMN])
        df.set_index(DATE_COLUMN, inplace=True)

        if period == '24h':
            df_res = df.resample('h').mean()
            wind_res = _wind_vector_resample(df, 'h')
            if not wind_res.empty:
                df_res = df_res.drop(columns=wind_res.columns.intersection(df_res.columns), errors='ignore')
                df_res = df_res.join(wind_res)
            if 'EVAPOR_MINUTE' in df.columns:
                df_res['EVAPOR_MINUTE'] = df['EVAPOR_MINUTE'].resample('h').apply(_last_valid_value)
            if 'RAIN' in df.columns:
                df_res['RAIN'] = df['RAIN'].resample('h').mean()
        elif period == '30d':
            df_res = df.drop(columns=['RADIATION'], errors='ignore').resample('d').mean()
            wind_res = _wind_vector_resample(df, 'd')
            if not wind_res.empty:
                df_res = df_res.drop(columns=wind_res.columns.intersection(df_res.columns), errors='ignore')
                df_res = df_res.join(wind_res)
            if 'RADIATION' in df.columns:
                rad = df['RADIATION'].resample('d').sum(min_count=1) * KWH_PER_M2_FROM_HOUR
                df_res = df_res.join(rad.rename('RADIATION'))
            if 'EVAPOR_MINUTE' in df.columns:
                df_res['EVAPOR_MINUTE'] = df['EVAPOR_MINUTE'].resample('d').mean()
            if 'RAIN' in df.columns:
                rain_day = df['RAIN'].resample('d').sum(min_count=1)
                df_res['RAIN'] = rain_day
        else:
            df_res = df.drop(columns=['RADIATION'], errors='ignore').resample('ME').mean()
            wind_res = _wind_vector_resample(df, 'ME')
            if not wind_res.empty:
                df_res = df_res.drop(columns=wind_res.columns.intersection(df_res.columns), errors='ignore')
                df_res = df_res.join(wind_res)
            if 'RADIATION' in df.columns:
                daily_rad = df['RADIATION'].resample('d').sum(min_count=1) * KWH_PER_M2_FROM_HOUR
                rad = daily_rad.resample('ME').sum(min_count=1)
                df_res = df_res.join(rad.rename('RADIATION'))
            if 'EVAPOR_MINUTE' in df.columns:
                daily_evap = df['EVAPOR_MINUTE'].resample('d').mean()
                monthly_evap = daily_evap.resample('ME').sum(min_count=1)
                df_res = df_res.drop(columns=['EVAPOR_MINUTE'], errors='ignore').join(
                    monthly_evap.rename('EVAPOR_MINUTE')
                )
            if not df_res.empty:
                df_res.index = df_res.index.to_period('M').to_timestamp()
            if 'RAIN' in df.columns:
                daily_rain = df['RAIN'].resample('d').sum(min_count=1)
                rain_month = daily_rain.resample('ME').sum(min_count=1)
                df_res = df_res.drop(columns=['RAIN'], errors='ignore').join(
                    rain_month.rename('RAIN')
                )

        df_res = df_res.dropna(how='all')
        df_res.reset_index(inplace=True)
        # Replace NaN values with None to ensure valid JSON serialization
        df_res = df_res.astype(object).where(pd.notnull(df_res), None)

        result = {col: df_res[col].tolist() for col in df_res.columns}
        result[DATE_COLUMN] = [ts.isoformat() for ts in result[DATE_COLUMN]]
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error in /graph_data endpoint: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/statistics')
@login_required
def statistics_page():
    return render_template('statistics.html')


def _format_dt(dt: datetime) -> str:
    return dt.strftime('%Y-%m-%d %H:%M')


def _min_max_with_time(df: pd.DataFrame, column: str):
    series = df[column].dropna()
    if series.empty:
        return None
    min_val = series.min()
    min_time = series.idxmin()
    max_val = series.max()
    max_time = series.idxmax()
    return {
        'min': float(min_val),
        'min_time': _format_dt(min_time),
        'max': float(max_val),
        'max_time': _format_dt(max_time),
    }


def format_number(val: float) -> str:
    sign = '-' if val < 0 else ''
    val = abs(val)
    whole, frac = f"{val:.1f}".split('.')
    groups = []
    while whole:
        groups.append(whole[-3:])
        whole = whole[:-3]
    whole_with_space = ' '.join(reversed(groups))
    return f"{sign}{whole_with_space},{frac}"
  
  
def _last_valid_value(series: pd.Series) -> float:
    values = pd.to_numeric(series, errors="coerce").dropna()
    if values.empty:
        return np.nan
    return float(values.iloc[-1])


def _dew_point(temp_c: pd.Series, rel_hum: pd.Series) -> pd.Series:
    """Calculate dew point temperature given air temperature and relative humidity."""
    temp_c = pd.to_numeric(temp_c, errors="coerce")
    rel_hum = pd.to_numeric(rel_hum, errors="coerce")
    a = 17.27
    b = 237.7
    # Humidity values at or below zero lead to invalid logarithms; treat them as missing
    safe_humidity = rel_hum.where(rel_hum > 0)
    with np.errstate(divide='ignore', invalid='ignore'):
        alpha = (a * temp_c / (b + temp_c)) + np.log(safe_humidity / 100.0)
    dew_point = (b * alpha) / (a - alpha)
    return dew_point.where(safe_humidity.notna())


def _wind_vector_resample(
    df: pd.DataFrame,
    freq: str,
    direction_column: str = "WIND_DIR",
    speed_columns: Optional[Sequence[str]] = None,
) -> pd.DataFrame:
    if speed_columns is None:
        speed_columns = [col for col in df.columns if col.startswith("WIND_SPEED")]

    available_speeds = [col for col in speed_columns if col in df.columns]
    include_direction = direction_column in df.columns

    resampled_groups = df.resample(freq)
    resampled_index = resampled_groups.mean().index

    if not available_speeds and not include_direction:
        return pd.DataFrame(index=resampled_index)

    rows = []
    for _, group in resampled_groups:
        row: Dict[str, float] = {}
        computed_direction = np.nan
        for speed_col in available_speeds:
            if include_direction:
                subset = group[[speed_col, direction_column]].copy()
                stats = calculate_wind_stats(
                    subset, speed_col=speed_col, dir_col=direction_column
                )
                row[speed_col] = stats.mean_speed_resultant
                if np.isnan(computed_direction) and not np.isnan(stats.mean_dir):
                    computed_direction = stats.mean_dir
            else:
                row[speed_col] = pd.to_numeric(group[speed_col], errors="coerce").mean()

        if include_direction:
            if np.isnan(computed_direction):
                direction_series = pd.to_numeric(
                    group[direction_column], errors="coerce"
                )
                direction_only_df = pd.DataFrame(
                    {
                        "_unit_speed": np.where(direction_series.notna(), 1.0, np.nan),
                        direction_column: direction_series,
                    }
                )
                stats = calculate_wind_stats(
                    direction_only_df, speed_col="_unit_speed", dir_col=direction_column
                )
                computed_direction = stats.mean_dir
            row[direction_column] = computed_direction

        rows.append(row)

    return pd.DataFrame(rows, index=resampled_index)


def _period_bounds(period: str):
    now = datetime.now()
    if period == 'today':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)
    elif period == 'month':
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end = (start + timedelta(days=32)).replace(day=1)
    elif period == 'year':
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end = start.replace(year=start.year + 1)
    else:
        start = None
        end = None
    return start, end


def _range_conditions(start: Optional[datetime], end: Optional[datetime]):
    conditions: List[str] = []
    params: List[datetime] = []
    if start is not None:
        conditions.append(f"{DATE_COLUMN} >= %s")
        params.append(start)
    if end is not None:
        conditions.append(f"{DATE_COLUMN} < %s")
        params.append(end)
    return conditions, params


def _compose_where_clause(conditions: List[str]) -> str:
    if not conditions:
        return ''
    return 'WHERE ' + ' AND '.join(conditions)


def _to_float(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


DEW_POINT_SQL_EXPR = (
    "CASE WHEN REL_HUM IS NULL OR REL_HUM <= 0 THEN NULL "
    "WHEN (17.27 - (((17.27 * T_AIR) / (237.7 + T_AIR)) + LN(REL_HUM / 100.0))) = 0 THEN NULL "
    "ELSE (237.7 * (((17.27 * T_AIR) / (237.7 + T_AIR)) + LN(REL_HUM / 100.0))) / "
    "(17.27 - (((17.27 * T_AIR) / (237.7 + T_AIR)) + LN(REL_HUM / 100.0))) END"
)


def _query_extrema(
    cursor,
    column: str,
    start: Optional[datetime],
    end: Optional[datetime],
    *,
    asc: bool = True,
    expression: Optional[str] = None,
    extra_columns: Optional[Sequence[str]] = None,
    filters: Optional[Sequence[str]] = None,
):
    extra_columns = list(extra_columns or [])
    conditions, params = _range_conditions(start, end)
    if expression is None:
        conditions.append(f"{column} IS NOT NULL")
    if filters:
        conditions.extend(filters)
    where_clause = _compose_where_clause(conditions)
    value_expr = expression or column
    select_parts = [f"{value_expr} AS value"] + extra_columns + [DATE_COLUMN]
    inner_query = (
        f"SELECT {', '.join(select_parts)} FROM {DB_TABLE_MIN} {where_clause}"
    )
    query = (
        f"SELECT * FROM ({inner_query}) AS stats "
        f"WHERE value IS NOT NULL ORDER BY value {'ASC' if asc else 'DESC'} LIMIT 1"
    )
    cursor.execute(query, params)
    row = cursor.fetchone()
    if not row:
        return None
    value = _to_float(row[0])
    if value is None:
        return None
    extras: Dict[str, Optional[float]] = {}
    for idx, col in enumerate(extra_columns):
        extras[col] = _to_float(row[1 + idx]) if isinstance(row[1 + idx], (Decimal, int, float)) else row[1 + idx]
    timestamp = row[1 + len(extra_columns)]
    if timestamp is None:
        return None
    return {"value": value, "timestamp": timestamp, "extras": extras}


def _query_sum(cursor, column: str, start: Optional[datetime], end: Optional[datetime]):
    conditions, params = _range_conditions(start, end)
    where_clause = _compose_where_clause(conditions)
    query = f"SELECT SUM({column}) FROM {DB_TABLE_MIN} {where_clause}"
    cursor.execute(query, params)
    row = cursor.fetchone()
    if not row:
        return None
    return _to_float(row[0])


def _query_evaporation_average(
    cursor, start: Optional[datetime], end: Optional[datetime]
) -> Optional[float]:
    conditions, params = _range_conditions(start, end)
    conditions.append("EVAPOR_MINUTE IS NOT NULL")
    where_clause = _compose_where_clause(conditions)
    query = f"SELECT AVG(EVAPOR_MINUTE) FROM {DB_TABLE_MIN} {where_clause}"
    cursor.execute(query, params)
    row = cursor.fetchone()
    if not row:
        return None
    return _to_float(row[0])


def _query_evaporation_daily_average_sum(
    cursor, start: Optional[datetime], end: Optional[datetime]
) -> Optional[float]:
    conditions, params = _range_conditions(start, end)
    conditions.append("EVAPOR_MINUTE IS NOT NULL")
    where_clause = _compose_where_clause(conditions)
    query = (
        f"SELECT DATE({DATE_COLUMN}) AS day, AVG(EVAPOR_MINUTE) AS avg_evap "
        f"FROM {DB_TABLE_MIN} {where_clause} GROUP BY day"
    )
    cursor.execute(query, params)
    rows = cursor.fetchall()
    if not rows:
        return None
    daily_values = [
        _to_float(row[1]) for row in rows if len(row) > 1 and _to_float(row[1]) is not None
    ]
    if not daily_values:
        return None
    return float(np.nansum(daily_values))


def _query_rain_total(cursor, start: Optional[datetime], end: Optional[datetime]):
    conditions, params = _range_conditions(start, end)
    where_clause = _compose_where_clause(conditions)
    query = f"SELECT SUM(RAIN) FROM {DB_TABLE} {where_clause}"
    cursor.execute(query, params)
    row = cursor.fetchone()
    if not row:
        return None
    return _to_float(row[0])


def _query_last_value(cursor, column: str, start: Optional[datetime], end: Optional[datetime]):
    conditions, params = _range_conditions(start, end)
    conditions.append(f"{column} IS NOT NULL")
    where_clause = _compose_where_clause(conditions)
    query = (
        f"SELECT {column}, {DATE_COLUMN} FROM {DB_TABLE_MIN} "
        f"{where_clause} ORDER BY {DATE_COLUMN} DESC LIMIT 1"
    )
    cursor.execute(query, params)
    row = cursor.fetchone()
    if not row:
        return None
    value = _to_float(row[0])
    timestamp = row[1]
    if value is None or timestamp is None:
        return None
    return {"value": value, "timestamp": timestamp}


def _query_daily_sum_extrema(
    cursor,
    column: str,
    start: Optional[datetime],
    end: Optional[datetime],
):
    conditions, params = _range_conditions(start, end)
    where_clause = _compose_where_clause(conditions)
    query = (
        f"SELECT DATE({DATE_COLUMN}) AS day, SUM(RAIN) AS total "
        f"FROM {DB_TABLE} {where_clause} GROUP BY day HAVING total IS NOT NULL "
        f"ORDER BY total DESC LIMIT 1"
    )
    cursor.execute(query, params)
    row = cursor.fetchone()
    if not row:
        return None
    day, total = row
    total_value = _to_float(total)
    if total_value is None or day is None:
        return None
    if isinstance(day, datetime):
        day_dt = day
    else:
        day_dt = datetime.combine(day, datetime.min.time())
    return {"value": total_value, "timestamp": day_dt}


def _build_stats(period: str, cursor):
    start, end = _period_bounds(period)

    entries: Dict[str, Dict[str, object]] = {}

    def add_entry(label: str, value):
        if value is None:
            return
        entries[label] = {"label": label, "value": value}

    temp_min = _query_extrema(cursor, 'T_AIR', start, end, asc=True)
    temp_max = _query_extrema(cursor, 'T_AIR', start, end, asc=False)
    if temp_min and temp_max:
        add_entry(
            "Температура",
            [
                f"мин {format_number(temp_min['value'])}°C ({_format_dt(temp_min['timestamp'])})",
                f"макс {format_number(temp_max['value'])}°C ({_format_dt(temp_max['timestamp'])})",
            ],
        )

    water_min = _query_extrema(cursor, 'T_WATER', start, end, asc=True)
    water_max = _query_extrema(cursor, 'T_WATER', start, end, asc=False)
    if water_min and water_max:
        add_entry(
            "Температура на водата",
            [
                f"мин {format_number(water_min['value'])}°C ({_format_dt(water_min['timestamp'])})",
                f"макс {format_number(water_max['value'])}°C ({_format_dt(water_max['timestamp'])})",
            ],
        )

    hum_min = _query_extrema(cursor, 'REL_HUM', start, end, asc=True)
    hum_max = _query_extrema(cursor, 'REL_HUM', start, end, asc=False)
    if hum_min and hum_max:
        add_entry(
            "Относителна влажност",
            [
                f"мин {format_number(hum_min['value'])}% ({_format_dt(hum_min['timestamp'])})",
                f"макс {format_number(hum_max['value'])}% ({_format_dt(hum_max['timestamp'])})",
            ],
        )

    dew_min = _query_extrema(
        cursor,
        'T_AIR',
        start,
        end,
        asc=True,
        expression=DEW_POINT_SQL_EXPR,
    )
    dew_max = _query_extrema(
        cursor,
        'T_AIR',
        start,
        end,
        asc=False,
        expression=DEW_POINT_SQL_EXPR,
    )
    if dew_min and dew_max:
        add_entry(
            "Точка на роса",
            [
                f"мин {format_number(dew_min['value'])}°C ({_format_dt(dew_min['timestamp'])})",
                f"макс {format_number(dew_max['value'])}°C ({_format_dt(dew_max['timestamp'])})",
            ],
        )

    press_rel_min = _query_extrema(cursor, 'P_REL', start, end, asc=True)
    press_rel_max = _query_extrema(cursor, 'P_REL', start, end, asc=False)
    if press_rel_min and press_rel_max:
        add_entry(
            "Относително налягане",
            [
                f"мин {format_number(press_rel_min['value'])} hPa ({_format_dt(press_rel_min['timestamp'])})",
                f"макс {format_number(press_rel_max['value'])} hPa ({_format_dt(press_rel_max['timestamp'])})",
            ],
        )

    press_abs_min = _query_extrema(cursor, 'P_ABS', start, end, asc=True)
    press_abs_max = _query_extrema(cursor, 'P_ABS', start, end, asc=False)
    if press_abs_min and press_abs_max:
        add_entry(
            "Абсолютно налягане",
            [
                f"мин {format_number(press_abs_min['value'])} hPa ({_format_dt(press_abs_min['timestamp'])})",
                f"макс {format_number(press_abs_max['value'])} hPa ({_format_dt(press_abs_max['timestamp'])})",
            ],
        )

    gust = _query_extrema(
        cursor,
        'WIND_GUST',
        start,
        end,
        asc=False,
        extra_columns=['WIND_DIR'],
    )
    if gust:
        direction = gust['extras'].get('WIND_DIR') if gust['extras'] else None
        if isinstance(direction, float) and np.isnan(direction):
            direction = None
        dir_text = f", посока {direction}" if direction is not None else ''
        add_entry(
            "Порив на вятъра",
            f"макс {format_number(gust['value'])} km/h{dir_text} ({_format_dt(gust['timestamp'])})",
        )

    rain_total = _query_rain_total(cursor, start, end)
    if rain_total:
        add_entry("Сума валежи", f"{format_number(rain_total)} mm")

    if period == 'today':
        evap_value = _query_evaporation_average(cursor, start, end)
        if evap_value is not None:
            add_entry("Изпарение", f"{format_number(evap_value)} mm")
    else:
        evap_sum = _query_evaporation_daily_average_sum(cursor, start, end)
        if evap_sum is not None:
            add_entry("Сума от изпарение", f"{format_number(evap_sum)} mm")

    if period != 'today':
        max_daily_rain = _query_daily_sum_extrema(cursor, 'RAIN', start, end)
        if max_daily_rain:
            label = "Максимално валежи за ден"
            add_entry(
                label,
                f"{format_number(max_daily_rain['value'])} mm ({_format_dt(max_daily_rain['timestamp'])})",
            )

        rain_intensity = _query_extrema(cursor, 'RAIN', start, end, asc=False)
        if rain_intensity:
            label = "Максимален интензитет"
            add_entry(
                label,
                f"{format_number(rain_intensity['value'])} mm/h ({_format_dt(rain_intensity['timestamp'])})",
            )

    radiation_max = _query_extrema(cursor, 'RADIATION', start, end, asc=False)
    if radiation_max:
        add_entry(
            "Слънчева радиация",
            f"макс {format_number(radiation_max['value'])} W/m² ({_format_dt(radiation_max['timestamp'])})",
        )

    radiation_sum = _query_sum(cursor, 'RADIATION', start, end)
    if radiation_sum is not None:
        energy = radiation_sum * KWH_PER_M2_FROM_MINUTE
        add_entry("Сума слънчева радиация", f"{format_number(energy)} kWh/m²")

    if period in ('month', 'year', 'all'):
        left_order = [
            "Температура",
            "Относителна влажност",
            "Относително налягане",
            "Абсолютно налягане",
            "Порив на вятъра",
            "Сума от изпарение",
        ]
        right_order = [
            "Температура на водата",
            "Точка на роса",
            "Сума валежи",
            "Максимален интензитет",
            "Максимално валежи за ден",
            "Слънчева радиация",
        ]
        ordered_labels = left_order + right_order
        result = [entries[label] for label in ordered_labels if label in entries]
        for label, entry in entries.items():
            if label not in ordered_labels:
                result.append(entry)
        return result

    return list(entries.values())


@app.route('/statistics_data')
@login_required
def statistics_data():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        data = {}
        for period in ('today', 'month', 'year', 'all'):
            data[period] = _build_stats(period, cursor)
        return jsonify(data)
    except Exception as e:
        logger.error(f"Error in /statistics_data endpoint: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if connection is not None:
            try:
                connection.close()
            except Exception:
                pass


@app.route('/report')
@login_required
def report_page():
    return render_template('report.html')


@app.route('/report_data', methods=['GET'])
@login_required
def report_data_endpoint():
    """Return daily statistics for a given month and year."""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    today = datetime.today().date()
    if not year or not month:
        return jsonify({})
    try:
        db_connection = get_db_connection()
        cursor = db_connection.cursor()

        # Discover available columns to avoid SQL errors if some are missing
        cursor.execute(f"SHOW COLUMNS FROM {DB_TABLE}")
        available = {row[0] for row in cursor.fetchall()}

        requested = [
            "T_AIR",
            "T_WATER",
            "REL_HUM",
            "P_REL",
            "P_ABS",
            "WIND_SPEED_1",
            "WIND_SPEED_2",
            "WIND_DIR",
            "RADIATION",
            "RAIN",
            "EVAPOR_MINUTE",
        ]
        cols = [c for c in requested if c in available]

        query = (
            f"SELECT {DATE_COLUMN}, {', '.join(cols)} FROM {DB_TABLE} "
            f"WHERE YEAR({DATE_COLUMN}) = %s AND MONTH({DATE_COLUMN}) = %s "
        )
        params = [year, month]
        if year == today.year and month == today.month:
            query += f"AND DATE({DATE_COLUMN}) < %s "
            params.append(today)
        query += f"ORDER BY {DATE_COLUMN} ASC"
        cursor.execute(query, params)
        data = cursor.fetchall()
        cursor.close()
        db_connection.close()

        df = pd.DataFrame(data, columns=[DATE_COLUMN] + cols)
        if df.empty:
            logger.error(f"No data returned for {year}-{month} from {DB_TABLE}")
            return jsonify({})

        # Convert numeric values and track columns that fail to parse
        for c in cols:
            raw = df[c].astype(str)
            df[c] = pd.to_numeric(raw.str.replace(",", "."), errors="coerce")
            if df[c].isna().all():
                logger.error(
                    "Column %s has no numeric data for %04d-%02d; sample raw values: %s",
                    c,
                    year,
                    month,
                    raw.head().tolist(),
                )

        df[DATE_COLUMN] = pd.to_datetime(df[DATE_COLUMN])
        df.set_index(DATE_COLUMN, inplace=True)
        df.sort_index(inplace=True)

        import calendar

        days_in_month = calendar.monthrange(year, month)[1]
        idx = pd.date_range(start=datetime(year, month, 1), periods=days_in_month, freq="D")

        combined = pd.DataFrame(index=idx)

        mean_cols = [
            c
            for c in [
                "T_AIR",
                "T_WATER",
                "REL_HUM",
                "P_REL",
                "P_ABS",
                "WIND_SPEED_1",
                "WIND_SPEED_2",
            ]
            if c in df.columns
        ]
        if mean_cols:
            combined = combined.join(df[mean_cols].resample("D").mean())

        wind_daily = _wind_vector_resample(df, "D")
        if not wind_daily.empty:
            combined = combined.drop(columns=wind_daily.columns.intersection(combined.columns), errors="ignore")
            combined = combined.join(wind_daily)

        if "RAIN" in df.columns:
            combined = combined.join(
                df["RAIN"].resample("D").sum(min_count=1).rename("RAIN")
            )

        if "EVAPOR_MINUTE" in df.columns:
            combined = combined.join(
                df["EVAPOR_MINUTE"].resample("D").mean().rename("EVAPOR_DAY")
            )

        if "RADIATION" in df.columns:
            rad_daily = df["RADIATION"].resample("D").sum(min_count=1) * KWH_PER_M2_FROM_HOUR
            combined = combined.join(rad_daily.rename("RADIATION"))

        at_14_cols = [c for c in ["T_AIR", "REL_HUM", "P_REL"] if c in df.columns]
        if at_14_cols:
            combined = combined.join(
                df.between_time("14:00", "14:00")[at_14_cols]
                .resample("D")
                .mean()
                .add_suffix("_14")
            )

        combined = combined.round(1)

        now = datetime.now()
        if year == now.year and month == now.month:
            today = now.date()
            if today in combined.index:
                combined.loc[today] = np.nan

        for col in combined.columns:
            if combined[col].isna().all():
                logger.error(
                    "No computed data for column %s in %04d-%02d after aggregation",
                    col,
                    year,
                    month,
                )

        result = {
            col: [None if pd.isna(v) else float(v) for v in combined[col]]
            for col in combined.columns
        }
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error in /report_data endpoint: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/plot', methods=['POST'])
def plot():
    global my_df
    try:
        # Get start and end dates from the frontend
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        category = request.form['category']

        # Add 1 hour to end_date
        end_date = pd.to_datetime(end_date) + pd.Timedelta(days=1)

        # Connect to the database
        db_connection = get_db_connection()

        # Fetch data from the database based on the date range
        cursor = db_connection.cursor()
        query = f"SELECT {DATE_COLUMN}, {', '.join(RAW_DATA_COLUMNS)} FROM {DB_TABLE} " \
                f"WHERE {DATE_COLUMN} >= %s AND {DATE_COLUMN} <= %s ORDER BY {DATE_COLUMN} ASC "
        cursor.execute(query, (start_date+' 00:00:00', end_date.strftime('%Y-%m-%d %H:%M:%S')))
        data = cursor.fetchall()

        # Close database connection
        cursor.close()
        db_connection.close()

        # Create a DataFrame from the fetched data with only DATE_COLUMN and consumption columns
        df = pd.DataFrame(data, columns=[DATE_COLUMN] + RAW_DATA_COLUMNS)

        # Convert DATE_COLUMN to datetime
        df[DATE_COLUMN] = pd.to_datetime(df[DATE_COLUMN])

        # Set DATE_COLUMN as index for proper time-based grouping
        df.set_index(DATE_COLUMN, inplace=True)
        df = add_calculated_columns(df)

        # Group data by category (hourly, daily, or monthly)
        if category == "hourly":
            df_grouped = df.groupby(pd.Grouper(freq='h')).mean()
        elif category == "daily":
            df_grouped = df.groupby(pd.Grouper(freq='D')).mean()
        elif category == "monthly":
            df_grouped = df.groupby(pd.Grouper(freq='ME')).mean()
        df_grouped = add_calculated_columns(df_grouped)
        df_grouped.reset_index(inplace=True)
        df_grouped.rename(columns={'index': DATE_COLUMN}, inplace=True)

        my_df = df_grouped
        columns_to_plot = [col for col in DATA_COLUMNS if col in my_df.columns]

        # Create Plotly Express figure for consumption
        fig1 = px.line(my_df, x=DATE_COLUMN, y=columns_to_plot, title="Всички данни")

        # Update axis titles
        fig1.update_xaxes(
            title_text="Дата",
            showspikes=True,
            spikecolor="#808080",
            spikethickness=1,
            spikedash="dot",
            spikemode="across",
            spikesnap="cursor",
        )
        fig1.update_yaxes(
            showspikes=True,
            spikecolor="#808080",
            spikethickness=1,
            spikedash="dot",
        )

        # Use the same title layout as fig3
        fig1.update_layout(
            hovermode="x unified",
            hoverdistance=100,
            spikedistance=-1,
            height=600,  # Adjust the height
            legend_title_text="Данни",
            title={
                'text': "Всички данни",
                'x': 0.5,  # Center the title
                'xanchor': 'center',  # Anchor the title at the center
                'yanchor': 'top'  # Anchor the title at the top
            },
            autosize=True  # Ensure the figure resizes dynamically

        )

        # Update legend names using Bulgarian column names
        bg_names_for_plot = [DATA_COLUMNS_BG[DATA_COLUMNS.index(c)] for c in columns_to_plot]
        for trace, data_col, data_col_name in zip(fig1.data, columns_to_plot, bg_names_for_plot):
            trace.name = data_col_name  # Set the trace name
            trace.hovertemplate = None  # Remove the default hover template

        # Update trace mode to markers+lines
        fig1.update_traces(mode="markers+lines", hovertemplate=None)

        # Wind Rose Plot for wind direction and speed
        wind_angle = my_df['WIND_DIR']
        wind_speed = my_df['WIND_SPEED_1']


        # Example wind speed and direction data
        wind_rose_fig = go.Figure(go.Scatterpolar(
            r=wind_speed,
            theta=wind_angle,
            mode='markers',
            marker=dict(size=8, color=wind_speed, colorscale='Viridis', showscale=True),
            name="Wind Rose"
        ))

        wind_rose_fig.update_layout(
            polar=dict(
                radialaxis=dict(visible=True),
                angularaxis=dict(
                                rotation=90,  # Start from 0° at the top (North)
                                direction="clockwise",
                                tickmode='array',
                                 tickvals=[0, 45, 90, 135, 180, 225, 270, 315],
                                 ticktext=['North', 'N-E', 'East', 'S-E', 'South', 'S-W', 'West', 'N-W'])
            ),
            title={
                'text': "Роза на ветровете",
                'x': 0.5,  # Center the title
                'xanchor': 'center',  # Anchor the title at the center
            },
            # height=600,  # Adjust the height
            # width=800,# Adjust the width
            autosize=True
        )
        # Convert plots to JSON format
        plot_json1 = fig1.to_json()
        wind_rose_json = wind_rose_fig.to_json()

        # Return plots in a JSON response
        return jsonify({
            'plot1': plot_json1,
            'wind_rose': wind_rose_json
        })
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        return str(e), 500


@app.route('/export_to_excel', methods=['POST'])
def export_to_excel():
    global my_df
    try:
        # Parse JSON data
        data = request.get_json()
        start_date = pd.to_datetime(data['start_date'])
        end_date = pd.to_datetime(data['end_date'])
        selected_traces = data.get('selected_traces', [])

        # Create translation dictionaries from lists
        column_map_latin_to_bg = dict(zip(DATA_COLUMNS, DATA_COLUMNS_BG))
        column_map_bg_to_latin = dict(zip(DATA_COLUMNS_BG, DATA_COLUMNS))

        # Translate selected traces from Bulgarian to Latin
        selected_traces_latin = [column_map_bg_to_latin.get(trace, trace) for trace in selected_traces]

        # Filter DataFrame
        if selected_traces_latin:
            filtered_df = my_df[['DateRef'] + selected_traces_latin]
        else:
            filtered_df = my_df

        # Translate column names to Bulgarian for export
        selected_columns_bg = ['Дата'] + [column_map_latin_to_bg.get(col, col) for col in selected_traces_latin]

        # Load Excel template
        wb = openpyxl.load_workbook(EXCEL_TEMPLATE_PATH)
        ws = wb.active

        # Write metadata
        ws['E8'] = start_date.strftime('%d-%m-%Y')
        ws['E9'] = end_date.strftime('%d-%m-%Y')

        # Write headers
        for col_idx, col_name in enumerate(selected_columns_bg, start=2):
            ws.cell(row=11, column=col_idx, value=col_name)

        # Write data
        for row_idx, (_, row) in enumerate(filtered_df.iterrows(), start=12):
            for col_idx, value in enumerate(row, start=2):
                if isinstance(value, str):
                    try:
                        value = float(value.replace(',', '.'))
                    except ValueError:
                        pass
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save and send file
        file_path = os.path.abspath('output.xlsx')
        wb.save(file_path)
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='output.xlsx'
        )
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        return jsonify({"error": "Failed to export data"}), 500


@app.route('/export2_to_excel', methods=['POST'])
def export2_to_excel():
    global my_df
    try:
        # Parse JSON data
        data = request.get_json()
        start_date = pd.to_datetime(data['start_date'])
        end_date = pd.to_datetime(data['end_date'])
        selected_traces = data.get('selected_traces', [])

        # Create translation dictionaries from lists
        column_map_latin_to_bg = dict(zip(DATA_COLUMNS, DATA_COLUMNS_BG))
        column_map_bg_to_latin = dict(zip(DATA_COLUMNS_BG, DATA_COLUMNS))

        # Translate selected traces from Bulgarian to Latin
        selected_traces_latin = [column_map_bg_to_latin.get(trace, trace) for trace in selected_traces]

        # Filter DataFrame
        if selected_traces_latin:
            filtered_df = my_df[['DateRef'] + selected_traces_latin]
        else:
            filtered_df = my_df

        # Translate column names to Bulgarian for export
        selected_columns_bg = ['Дата'] + [column_map_latin_to_bg.get(col, col) for col in selected_traces_latin]

        # Load Excel template
        wb = openpyxl.load_workbook(EXCEL_TEMPLATE_PATH)
        ws = wb.active

        # Write metadata
        ws['E8'] = start_date.strftime('%Y-%m-%d')
        ws['E9'] = end_date.strftime('%Y-%m-%d')

        # Write headers
        for col_idx, col_name in enumerate(selected_columns_bg, start=2):
            ws.cell(row=11, column=col_idx, value=col_name)

        # Write data
        for row_idx, (_, row) in enumerate(filtered_df.iterrows(), start=12):
            for col_idx, value in enumerate(row, start=2):
                if isinstance(value, str):
                    try:
                        value = float(value.replace(',', '.'))
                    except ValueError:
                        pass
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save and send file
        file_path = os.path.abspath('output.xlsx')
        wb.save(file_path)
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='output.xlsx'
        )
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        return jsonify({"error": "Failed to export data"}), 500


def run_insert_missing_data():
    try:
        insertMissingDataFromCSV.main()
    except Exception as e:
        logger.error(f"Error updating data from CSV: {e}")


@app.route('/insert_missing_data', methods=['POST'])
def insert_missing_data():
    try:
        run_insert_missing_data()
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/moment_data', methods=['GET'])
def moment_data():
    global df_last_min_values

    def format_date(record):
        try:
            if 'DateRef' in record and record['DateRef']:
                if isinstance(record['DateRef'], pd.Timestamp):
                    record['DateRef'] = record['DateRef'].strftime('%d.%m.%Y %H:%M')
                else:
                    record['DateRef'] = datetime.strptime(record['DateRef'], '%a, %d %b %Y %H:%M:%S %Z').strftime('%d.%m.%Y %H:%M')
        except Exception as e:
            print(f"Error formatting date: {e}")

    def prepare_data(df):
        return df.reset_index(drop=True).to_dict('records') if not df.empty else []

    try:
        update_type = request.args.get("update_type", "last_minute")
        if update_type != "last_minute":
            return jsonify({"error": "Invalid update_type"}), 400

        df_last_min = df_last_min_values.copy()
        df_last_min.fillna(0, inplace=True)
        df_last_min.round(1)

        min_values_data = prepare_data(df_last_min)
        for record in min_values_data:
            format_date(record)

        alarms_info = get_active_alarms()
        ftp_status = getattr(insertMissingDataFromCSV, 'last_ftp_status', None)
        if isinstance(ftp_status, np.bool_):
            ftp_status = bool(ftp_status)

        return jsonify({
            "min_values_data": min_values_data,
            "columns_values": DATA_COLUMNS,
            "columns_bg": DATA_COLUMNS_BG,
            "columns_units": DATA_COLUMNS_UNITS,
            "alarms": alarms_info.get('items', []),
            "alarms_timestamp": alarms_info.get('timestamp'),
            "ftp_connection_ok": ftp_status,
        })

    except Exception as e:
        print(f"Error in /moment_data endpoint: {str(e)}")
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# Initialize BackgroundScheduler
scheduler = BackgroundScheduler()

# Job to compute hourly averages
def run_hourly_mean():
    now = datetime.now()
    mean_1h.mean_1h(now, now)

# Function to start the scheduler
def start_scheduler():
    scheduler.add_job(run_insert_missing_data, 'cron', hour='*', minute='*', second='1')
    scheduler.add_job(run_hourly_mean, 'cron', minute='0', second='30')
    scheduler.start()
    print("Scheduler started in a separate thread...")

# Start scheduler in a separate thread
threading.Thread(target=start_scheduler, daemon=True).start()

# Stop the scheduler gracefully when the app stops
def shutdown_scheduler(sender, **extra):
    if scheduler.running:
        scheduler.shutdown()
        print("Scheduler has been shut down gracefully.")

# Attach the shutdown function to the 'got_request_exception' signal
got_request_exception.connect(shutdown_scheduler, app)

if init == 0:
    run_insert_missing_data()

if __name__ == '__main__':
    #insertMissingDataFromCSV.main()
    #app.run(host='0.0.0.0', port=5010, debug=False)
    # uncomment this to start in non production
    serve(app, host='0.0.0.0', port=5010)

